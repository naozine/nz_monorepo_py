from pathlib import Path
from typing import Union, List, Dict, Any
import sys

try:
    # openpyxl is the de-facto library for .xlsx read/write
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill
except Exception as e:  # pragma: no cover
    raise RuntimeError("openpyxl が必要です。`pip install openpyxl` を実行してください。") from e

try:
    import yaml  # PyYAML
except Exception as e:  # pragma: no cover
    yaml = None

try:
    from main import ReportDataPreparator, ReportConfig, ProcessedData
except Exception as e:  # pragma: no cover
    print(f"警告: main.py からのインポートに失敗しました: {e}")
    print("survey_data 機能は利用できません。")
    ReportDataPreparator = None


def fill_ac14(template_path: Union[str, Path], output_path: Union[str, Path], value: str = "xxx") -> Path:
    """
    report_template.xlsx を読み込み、シート p1 の AC14 セルを指定値に更新して保存します。

    :param template_path: 入力テンプレートのパス（.xlsx）
    :param output_path: 出力先ファイルのパス（.xlsx）
    :param value: AC14 に書き込む値（デフォルト: "xxx"）
    :return: 作成したファイルの Path
    """
    tpath = Path(template_path)
    if not tpath.exists():
        raise FileNotFoundError(f"テンプレートが見つかりません: {tpath}")

    wb = load_workbook(tpath)

    if "p1" not in wb.sheetnames:
        raise KeyError("テンプレートにシート 'p1' が見つかりません")

    ws = wb["p1"]
    # 値を書き込み、変更有無に応じて背景色を付与（変更: クリーム／未変更: 薄い水色）
    write_with_cream(ws, "AC14", value)

    out = Path(output_path)
    # 出力ディレクトリが存在しない場合に備える
    out.parent.mkdir(parents=True, exist_ok=True)

    wb.save(out)
    return out


def get_survey_data_value(survey_data_type: str, processed_data: 'ProcessedData' = None, excel_path: Union[str, Path] = "survey.xlsx", question: int = None, choices: list = None, choice_mapping = None, class_type: str = None) -> Any:
    """
    survey_data_type に応じて、main.py の集計データから値を取得します。
    
    :param survey_data_type: 取得するデータタイプ
        - "total_responses": 総回答者数
        - "invalid_responses": 無効回答者数（未就学児等）
        - "effective_responses": 有効回答者数
        - "pick_count:q=<番号>;count=<回数>": 指定設問で指定回数の選択肢を選んだ人数
        - "multiple": 複数選択肢をすべて選択した回答者数
    :param processed_data: 事前に処理済みのデータ（パフォーマンス最適化用）
    :param excel_path: アンケートExcelファイルのパス（processed_data未指定時のみ使用）
    :return: 対応する値
    """
    if ReportDataPreparator is None:
        raise RuntimeError("main.py からのインポートが失敗しているため、survey_data 機能は利用できません。")
    
    try:
        # 事前処理済みデータが提供されていない場合のみ新規作成
        if processed_data is None:
            config = ReportConfig()
            preparator = ReportDataPreparator(config)
            processed_data = preparator.prepare_data(Path(excel_path))
        
        if survey_data_type == "total_responses":
            return processed_data.n_total + processed_data.n_preschool
        elif survey_data_type == "invalid_responses":
            return processed_data.n_preschool
        elif survey_data_type == "effective_responses":
            return processed_data.n_total
        elif survey_data_type.startswith("pick_count"):
            # 記法: "pick_count:q=<番号>;count=<回数>"
            params_str = ""
            parts = survey_data_type.split(":", 1)
            if len(parts) == 2:
                params_str = parts[1]
            # セミコロン区切りの key=value
            params: Dict[str, str] = {}
            if params_str:
                for token in params_str.split(";"):
                    if "=" in token:
                        k, v = token.split("=", 1)
                        params[k.strip().lower()] = v.strip()
            # 必須の2要素
            q_str = params.get("q") or params.get("question")
            count_str = params.get("count")
            if not q_str or not count_str:
                raise ValueError("pick_count には q(またはquestion) / count を指定してください。例: pick_count:q=6;count=2")
            try:
                q_idx = int(q_str)
                target_count = int(count_str)
            except Exception:
                raise ValueError(f"pick_count の q と count は整数で指定してください（q: {q_str}, count: {count_str}）")
            
            # 設問列名の解決（1開始）
            questions = processed_data.question_columns
            if q_idx < 1 or q_idx > len(questions):
                raise IndexError(f"pick_count: question 番号が範囲外です（1〜{len(questions)}）。指定: {q_idx}")
            qcol = questions[q_idx - 1]
            if qcol not in processed_data.df_effective.columns:
                raise RuntimeError(f"指定の設問列が見つかりません: {qcol}")
            
            # 各回答者の選択肢数をカウント
            df = processed_data.df_effective
            def count_choices_in_cell(cell_value) -> int:
                if cell_value is None:
                    return 0
                try:
                    import pandas as _pd
                    if _pd.isna(cell_value):
                        return 0
                except:
                    pass
                cell_str = str(cell_value).strip()
                if not cell_str:
                    return 0
                # 改行区切りで分割
                import re as _re
                parts = _re.split(r"[\r\n]+", cell_str)
                choices = [p.strip() for p in parts if p.strip()]
                return len(choices)
            
            # 指定回数の選択肢を選んだ人数をカウント
            choice_counts = df[qcol].apply(count_choices_in_cell)
            return int((choice_counts == target_count).sum())
        elif survey_data_type == "multiple":
            # 複数選択肢をすべて選択した回答者数
            if not question or not choices:
                raise ValueError("survey_data_type 'multiple' には question と choices が必要です。")
            
            # 設問列名の解決（1開始）
            questions = processed_data.question_columns
            if question < 1 or question > len(questions):
                raise IndexError(f"question 番号が範囲外です（1〜{len(questions)}）。指定: {question}")
            qcol = questions[question - 1]
            if qcol not in processed_data.df_effective.columns:
                raise RuntimeError(f"指定の設問列が見つかりません: {qcol}")
            
            # 選択肢マッピング解決関数（既存ロジックを再利用）
            def resolve_choice_with_mapping(yaml_choice: str, choice_mapping, actual_choices: set, yaml_choices: list = None) -> str:
                # 1. マッピング定義チェック
                if choice_mapping:
                    mapped_choice = None
                    
                    # 辞書形式のmapping
                    if isinstance(choice_mapping, dict) and yaml_choice in choice_mapping:
                        mapped_choice = choice_mapping[yaml_choice]
                    
                    # リスト形式のmapping（yaml_choicesとの順序対応）
                    elif isinstance(choice_mapping, list) and yaml_choices:
                        try:
                            choice_index = yaml_choices.index(yaml_choice)
                            if choice_index < len(choice_mapping):
                                mapped_choice = choice_mapping[choice_index]
                        except ValueError:
                            pass
                    
                    if mapped_choice:
                        if mapped_choice in actual_choices:
                            return mapped_choice
                
                # 2. 従来の処理（完全一致 → 部分一致）
                if yaml_choice in actual_choices:
                    return yaml_choice
                
                for actual_choice in actual_choices:
                    if yaml_choice in actual_choice:
                        return actual_choice
                
                return None
            
            # 複数選択肢の同時選択をチェック
            def contains_all_choices(cell_value):
                if cell_value is None:
                    return False
                try:
                    import pandas as _pd
                    if _pd.isna(cell_value):
                        return False
                except:
                    pass
                
                # セル値を選択肢に分解
                cell_str = str(cell_value).strip()
                if not cell_str:
                    return False
                
                import re as _re
                parts = _re.split(r"[\r\n]+", cell_str)
                choices_in_cell = [p.strip() for p in parts if p.strip()]
                choices_set = set(choices_in_cell)
                
                # すべての指定選択肢が含まれているかチェック（早期終了最適化）
                for required_choice in choices:
                    matched_choice = resolve_choice_with_mapping(required_choice, choice_mapping, choices_set, choices)
                    if matched_choice is None:
                        return False  # 1つでも見つからなければ即座にFalse
                
                return True  # すべて見つかった
            
            # データフィルタリングとカウント
            df = processed_data.df_effective
            if class_type and class_type.lower() == "grade":
                # 学年別の集計（9要素のリストを返す）
                order = ["小1", "小2", "小3", "小4", "小5", "小6", "中1", "中2", "中3"]
                if "grade_2024" not in df.columns:
                    raise RuntimeError("必要な列 'grade_2024' が見つかりません。")
                
                result = []
                for grade in order:
                    grade_df = df[df["grade_2024"] == grade]
                    mask = grade_df[qcol].apply(contains_all_choices)
                    count = int(mask.sum())
                    result.append(count)
                return result
                
            elif class_type and class_type.lower() == "region":
                # 地域別の集計（6要素のリストを返す）
                order = ["東京23区", "三多摩島しょ", "埼玉県", "神奈川県", "千葉県", "その他"]
                if "region_bucket" not in df.columns:
                    raise RuntimeError("必要な列 'region_bucket' が見つかりません。")
                
                result = []
                for region in order:
                    region_df = df[df["region_bucket"] == region]
                    mask = region_df[qcol].apply(contains_all_choices)
                    count = int(mask.sum())
                    result.append(count)
                return result
                
            else:
                # 全体の集計（単一値を返す）
                mask = df[qcol].apply(contains_all_choices)
                return int(mask.sum())
        else:
            raise ValueError(f"不明な survey_data_type: {survey_data_type}")
            
    except Exception as e:
        raise RuntimeError(f"集計データの取得に失敗しました: {e}")


def get_survey_data_series(series_type: str, processed_data: 'ProcessedData' = None, excel_path: Union[str, Path] = "survey.xlsx", choice_mapping = None, yaml_choices: list = None, select_count: int = None) -> List[Union[int, float]]:
    """
    指定のシリーズ種別に応じて、人数配列を返します。
    
    サポートする series_type:
      - "elementary_boys": 小1〜小6の男子の回答者数（6個）
      - "elementary_girls": 小1〜小6の女子の回答者数（6個）
      - "junior_boys": 中1〜中3の男子の回答者数（3個）
      - "junior_girls": 中1〜中3の女子の回答者数（3個）
      - "region_grades:<地域名>": 指定地域の小1〜中3の回答者数（9個）
        例: "region_grades:東京23区" / "region_grades:三多摩島しょ" / "region_grades:埼玉県" など
      - "responses:q=<番号>;choice=<選択肢>;class=<grade|region>": 指定設問の特定選択肢の回答数を、学年または地域の順で返す
        例: "responses:q=1;choice=知っている;class=grade"
      - "ratios:q=<番号>;choices=<選択肢リスト>;class=<total|grade|region>": 指定設問の選択肢の回答割合（小数）を返す
        例: "ratios:q=1;choices=知っている,知らない;class=total"
    
    :param processed_data: 事前に処理済みのデータ（パフォーマンス最適化用）
    :param excel_path: アンケートExcelファイルのパス（processed_data未指定時のみ使用）
    :return: 人数リスト（responses）または割合リスト（ratios）（series_typeに応じた順序）
    """
    if ReportDataPreparator is None:
        raise RuntimeError("main.py からのインポートが失敗しているため、survey_series 機能は利用できません。")

    try:
        # 事前処理済みデータが提供されていない場合のみ新規作成
        if processed_data is None:
            config = ReportConfig()
            preparator = ReportDataPreparator(config)
            processed_data = preparator.prepare_data(Path(excel_path))
        df = processed_data.df_effective

        # 地域名の正規化関数
        def normalize_region(name: str) -> str:
            nm = str(name).strip()
            if nm == "東京都下":
                return "三多摩島しょ"
            return nm

        # 新機能: 設問×選択肢の回答数シリーズ（学年または地域の順）
        if series_type.startswith("responses"):
            # 記法: "responses:q=<番号>;choice=<選択肢>;class=<grade|region>"
            # パーズ
            params_str = ""
            parts = series_type.split(":", 1)
            if len(parts) == 2:
                params_str = parts[1]
            # セミコロン区切りの key=value
            params: Dict[str, str] = {}
            if params_str:
                for token in params_str.split(";"):
                    if "=" in token:
                        k, v = token.split("=", 1)
                        params[k.strip().lower()] = v.strip()
            # 必須の3要素
            q_str = params.get("q") or params.get("question")
            choice = params.get("choice")
            class_type = params.get("class")
            if not q_str or not choice or not class_type:
                raise ValueError("responses シリーズには q(またはquestion) / choice / class を指定してください。例: responses:q=1;choice=知っている;class=grade")
            try:
                q_idx = int(q_str)
            except Exception:
                raise ValueError(f"responses の q は整数で指定してください（指定値: {q_str}）")
            class_type = class_type.lower()
            if class_type not in ("grade", "region", "total"):
                raise ValueError("responses の class は 'grade', 'region', または 'total' を指定してください。")

            # 設問列名の解決（1開始）
            questions = processed_data.question_columns
            if q_idx < 1 or q_idx > len(questions):
                raise IndexError(f"responses: question 番号が範囲外です（1〜{len(questions)}）。指定: {q_idx}")
            qcol = questions[q_idx - 1]
            if qcol not in df.columns:
                raise RuntimeError(f"指定の設問列が見つかりません: {qcol}")

            # 選択肢マッピング解決関数
            def resolve_choice_with_mapping(yaml_choice: str, choice_mapping, actual_choices: set, yaml_choices: list = None) -> str:
                """
                マッピング優先の選択肢解決
                choice_mappingは辞書またはリスト形式に対応
                1. choice_mappingに定義があれば、マッピング先で完全一致検索
                2. マッピングがなければ従来の部分一致処理
                """
                
                # 1. マッピング定義チェック
                if choice_mapping:
                    mapped_choice = None
                    
                    # 辞書形式のmapping
                    if isinstance(choice_mapping, dict) and yaml_choice in choice_mapping:
                        mapped_choice = choice_mapping[yaml_choice]
                    
                    # リスト形式のmapping（yaml_choicesとの順序対応）
                    elif isinstance(choice_mapping, list) and yaml_choices:
                        try:
                            choice_index = yaml_choices.index(yaml_choice)
                            if choice_index < len(choice_mapping):
                                mapped_choice = choice_mapping[choice_index]
                        except ValueError:
                            pass  # yaml_choice が yaml_choices に見つからない場合
                    
                    if mapped_choice:
                        if mapped_choice in actual_choices:
                            return mapped_choice
                        # マッピング先が見つからない場合（通常の動作 - セルごとに1つの値のみ含まれる）
                
                # 2. 従来の処理（完全一致 → 部分一致）
                if yaml_choice in actual_choices:
                    return yaml_choice
                
                for actual_choice in actual_choices:
                    if yaml_choice in actual_choice:
                        return actual_choice
                
                return None

            # 効率的な選択肢カウント関数（パンダスベクトル化）
            def efficient_choice_counting(df_subset, qcol: str, choice: str, choice_mapping, class_col: str, class_order: list, yaml_choices: list = None, select_count: int = None) -> List[int]:
                """
                高速なベクトル化処理で選択肢をカウント
                部分一致にも対応
                select_count が指定された場合、その選択数の回答者のみを対象とする
                """
                def contains_choice(cell_value):
                    if cell_value is None:
                        return False
                    try:
                        import pandas as _pd
                        if _pd.isna(cell_value):
                            return False
                    except:
                        pass
                    
                    # セル値を選択肢に分解
                    cell_str = str(cell_value).strip()
                    if not cell_str:
                        return False
                    
                    import re as _re
                    parts = _re.split(r"[\r\n]+", cell_str)
                    choices_in_cell = [p.strip() for p in parts if p.strip()]
                    
                    # マッピング対応の選択肢解決
                    matched_choice = resolve_choice_with_mapping(choice, choice_mapping, set(choices_in_cell), yaml_choices)
                    return matched_choice is not None
                
                # select_count によるフィルタリング
                if select_count is not None:
                    def count_choices_in_cell(cell_value) -> int:
                        if cell_value is None:
                            return 0
                        try:
                            import pandas as _pd
                            if _pd.isna(cell_value):
                                return 0
                        except:
                            pass
                        cell_str = str(cell_value).strip()
                        if not cell_str:
                            return 0
                        import re as _re
                        parts = _re.split(r"[\r\n]+", cell_str)
                        choices = [p.strip() for p in parts if p.strip()]
                        return len(choices)
                    
                    # 指定された選択数の回答者のみをフィルタリング
                    choice_counts = df_subset[qcol].apply(count_choices_in_cell)
                    df_subset = df_subset[choice_counts == select_count]
                
                # より効率的: copyを避けてSeriesで処理
                choice_mask = df_subset[qcol].apply(contains_choice)
                
                # マスクを使って該当行のみを集計
                valid_rows = df_subset[choice_mask]
                
                if len(valid_rows) == 0:
                    return [0] * len(class_order)
                
                # グループ化して一括集計
                result = valid_rows.groupby(class_col).size().reindex(class_order, fill_value=0)
                
                return [int(x) for x in result.tolist()]

            # セルから選択肢集合を得る関数（改行区切りにも対応）- 後方互換用
            def cell_to_set(val) -> set:
                if val is None:
                    return set()
                try:
                    import pandas as _pd  # 局所importでNaN検出
                    if _pd.isna(val):
                        return set()
                except Exception:
                    pass
                s = str(val).strip()
                if not s:
                    return set()
                import re as _re
                parts = _re.split(r"[\r\n]+", s)
                return {p.strip() for p in parts if p.strip()}

            # 高速集計（ベクトル化処理）
            if class_type == "grade":
                order = ["小1", "小2", "小3", "小4", "小5", "小6", "中1", "中2", "中3"]
                if "grade_2024" not in df.columns:
                    raise RuntimeError("必要な列 'grade_2024' が見つかりません。")
                return efficient_choice_counting(df, qcol, choice, choice_mapping, "grade_2024", order, yaml_choices, select_count)
            elif class_type == "region":
                order = ["東京23区", "三多摩島しょ", "埼玉県", "神奈川県", "千葉県", "その他"]
                if "region_bucket" not in df.columns:
                    raise RuntimeError("必要な列 'region_bucket' が見つかりません。")
                return efficient_choice_counting(df, qcol, choice, choice_mapping, "region_bucket", order, yaml_choices, select_count)
            else:  # total
                # 全体の集計: 単一の値を返す（リスト形式で1要素）
                def contains_choice(cell_value):
                    if cell_value is None:
                        return False
                    try:
                        import pandas as _pd
                        if _pd.isna(cell_value):
                            return False
                    except:
                        pass
                    
                    # セル値を選択肢に分解
                    cell_str = str(cell_value).strip()
                    if not cell_str:
                        return False
                    
                    import re as _re
                    parts = _re.split(r"[\r\n]+", cell_str)
                    choices_in_cell = [p.strip() for p in parts if p.strip()]
                    
                    # マッピング対応の選択肢解決
                    matched_choice = resolve_choice_with_mapping(choice, choice_mapping, set(choices_in_cell), yaml_choices)
                    return matched_choice is not None
                
                # select_count によるフィルタリング
                df_filtered = df
                if select_count is not None:
                    def count_choices_in_cell(cell_value) -> int:
                        if cell_value is None:
                            return 0
                        try:
                            import pandas as _pd
                            if _pd.isna(cell_value):
                                return 0
                        except:
                            pass
                        cell_str = str(cell_value).strip()
                        if not cell_str:
                            return 0
                        import re as _re
                        parts = _re.split(r"[\r\n]+", cell_str)
                        choices = [p.strip() for p in parts if p.strip()]
                        return len(choices)
                    
                    # 指定された選択数の回答者のみをフィルタリング
                    choice_counts = df[qcol].apply(count_choices_in_cell)
                    df_filtered = df[choice_counts == select_count]
                
                # 選択肢マッチング
                choice_mask = df_filtered[qcol].apply(contains_choice)
                count = int(choice_mask.sum())
                return [count]

        # 新機能: 設問×選択肢の回答割合シリーズ（全体・学年・地域別）
        if series_type.startswith("ratios"):
            # 記法: "ratios:q=<番号>;choices=<選択肢1>,<選択肢2>;class=<total|grade|region>"
            # パーズ
            params_str = ""
            parts = series_type.split(":", 1)
            if len(parts) == 2:
                params_str = parts[1]
            # セミコロン区切りの key=value
            params: Dict[str, str] = {}
            if params_str:
                for token in params_str.split(";"):
                    if "=" in token:
                        k, v = token.split("=", 1)
                        params[k.strip().lower()] = v.strip()
            
            # 必須の3要素
            q_str = params.get("q") or params.get("question")
            choices_str = params.get("choices")
            class_type = params.get("class")
            if not q_str or not choices_str or not class_type:
                raise ValueError("ratios シリーズには q(またはquestion) / choices / class を指定してください。例: ratios:q=1;choices=知っている,知らない;class=total")
            
            try:
                q_idx = int(q_str)
            except Exception:
                raise ValueError(f"ratios の q は整数で指定してください（指定値: {q_str}）")
            
            class_type = class_type.lower()
            if class_type not in ("total", "grade", "region"):
                raise ValueError("ratios の class は 'total', 'grade', または 'region' を指定してください。")

            # 選択肢のパース（カンマ区切り）
            choices = [c.strip() for c in choices_str.split(",") if c.strip()]
            if not choices:
                raise ValueError("ratios の choices は1つ以上の選択肢をカンマ区切りで指定してください。")

            # 設問列名の解決（1開始）
            questions = processed_data.question_columns
            if q_idx < 1 or q_idx > len(questions):
                raise IndexError(f"ratios: question 番号が範囲外です（1〜{len(questions)}）。指定: {q_idx}")
            qcol = questions[q_idx - 1]
            if qcol not in df.columns:
                raise RuntimeError(f"指定の設問列が見つかりません: {qcol}")

            # 選択肢マッピング解決関数（既存のものを再利用）
            def resolve_choice_with_mapping(yaml_choice: str, choice_mapping, actual_choices: set, yaml_choices: list = None) -> str:
                """
                マッピング優先の選択肢解決
                choice_mappingは辞書またはリスト形式に対応
                1. choice_mappingに定義があれば、マッピング先で完全一致検索
                2. マッピングがなければ従来の部分一致処理
                """
                
                # 1. マッピング定義チェック
                if choice_mapping:
                    mapped_choice = None
                    
                    # 辞書形式のmapping
                    if isinstance(choice_mapping, dict) and yaml_choice in choice_mapping:
                        mapped_choice = choice_mapping[yaml_choice]
                    
                    # リスト形式のmapping（yaml_choicesとの順序対応）
                    elif isinstance(choice_mapping, list) and yaml_choices:
                        try:
                            choice_index = yaml_choices.index(yaml_choice)
                            if choice_index < len(choice_mapping):
                                mapped_choice = choice_mapping[choice_index]
                        except ValueError:
                            pass  # yaml_choice が yaml_choices に見つからない場合
                    
                    if mapped_choice:
                        if mapped_choice in actual_choices:
                            return mapped_choice
                        # マッピング先が見つからない場合（通常の動作 - セルごとに1つの値のみ含まれる）
                
                # 2. 従来の処理（完全一致 → 部分一致）
                if yaml_choice in actual_choices:
                    return yaml_choice
                
                for actual_choice in actual_choices:
                    if yaml_choice in actual_choice:
                        return actual_choice
                
                return None

            # 割合計算用関数
            def calculate_ratios_for_choices(df_subset, qcol: str, choices: List[str], choice_mapping, yaml_choices: list = None) -> List[float]:
                """
                指定された選択肢リストの回答割合を計算
                """
                total_responses = len(df_subset)
                if total_responses == 0:
                    return [0.0] * len(choices)
                
                ratios = []
                for choice in choices:
                    def contains_choice(cell_value):
                        if cell_value is None:
                            return False
                        try:
                            import pandas as _pd
                            if _pd.isna(cell_value):
                                return False
                        except:
                            pass
                        
                        # セル値を選択肢に分解
                        cell_str = str(cell_value).strip()
                        if not cell_str:
                            return False
                        
                        import re as _re
                        parts = _re.split(r"[\r\n]+", cell_str)
                        choices_in_cell = [p.strip() for p in parts if p.strip()]
                        
                        # マッピング対応の選択肢解決
                        matched_choice = resolve_choice_with_mapping(choice, choice_mapping, set(choices_in_cell), yaml_choices)
                        return matched_choice is not None
                    
                    # 該当する回答数をカウント
                    choice_count = df_subset[qcol].apply(contains_choice).sum()
                    ratio = float(choice_count) / float(total_responses)
                    ratios.append(ratio)
                
                return ratios

            # クラス別の処理
            if class_type == "total":
                # 全体の割合を返す
                return calculate_ratios_for_choices(df, qcol, choices, choice_mapping, yaml_choices)
            
            elif class_type == "grade":
                # 学年別の割合を返す（各選択肢 × 各学年の二次元配列を一次元化）
                order = ["小1", "小2", "小3", "小4", "小5", "小6", "中1", "中2", "中3"]
                if "grade_2024" not in df.columns:
                    raise RuntimeError("必要な列 'grade_2024' が見つかりません。")
                
                result = []
                for grade in order:
                    grade_df = df[df["grade_2024"] == grade]
                    grade_ratios = calculate_ratios_for_choices(grade_df, qcol, choices, choice_mapping, yaml_choices)
                    result.extend(grade_ratios)
                
                return result
            
            else:  # region
                # 地域別の割合を返す（各選択肢 × 各地域の二次元配列を一次元化）
                order = ["東京23区", "三多摩島しょ", "埼玉県", "神奈川県", "千葉県", "その他"]
                if "region_bucket" not in df.columns:
                    raise RuntimeError("必要な列 'region_bucket' が見つかりません。")
                
                result = []
                for region in order:
                    region_df = df[df["region_bucket"] == region]
                    region_ratios = calculate_ratios_for_choices(region_df, qcol, choices, choice_mapping, yaml_choices)
                    result.extend(region_ratios)
                
                return result

        # 地域シリーズ（region_grades:...）の特別処理
        if series_type.startswith("region_grades"):
            # 記法1: "region_grades:地域名"
            region_name = None
            parts = series_type.split(":", 1)
            if len(parts) == 2 and parts[1].strip():
                region_name = parts[1].strip()
            # 記法2: YAML 側で region キーを使う場合は fill_from_yaml 側で地域名を補完して再呼び出しする想定
            # ここでは parts に地域名があればそれを使う
            if not region_name:
                raise ValueError("region_grades の地域名が指定されていません。'region_grades:東京23区' のように指定してください。")
            reg = normalize_region(region_name)
            allowed = {"東京23区", "三多摩島しょ", "埼玉県", "神奈川県", "千葉県", "その他"}
            if reg not in allowed:
                raise ValueError(f"不明な地域名: {region_name}（許可: 東京23区, 三多摩島しょ, 埼玉県, 神奈川県, 千葉県, その他）")

            # 必要な列確認
            for col in ["region_bucket", "grade_2024"]:
                if col not in df.columns:
                    raise RuntimeError(f"必要な列 '{col}' が見つかりません。Excelや前処理の仕様をご確認ください。")

            grades = ["小1", "小2", "小3", "小4", "小5", "小6", "中1", "中2", "中3"]
            counts: List[int] = []
            mask_reg = (df["region_bucket"] == reg)
            for g in grades:
                counts.append(int(((df["grade_2024"] == g) & mask_reg).sum()))
            return counts

        # 既存：性別×学年シリーズ
        series_def = {
            "elementary_boys": {
                "level": "小学校",
                "gender": "男性",
                "grades": ["小1", "小2", "小3", "小4", "小5", "小6"],
            },
            "elementary_girls": {
                "level": "小学校",
                "gender": "女性",
                "grades": ["小1", "小2", "小3", "小4", "小5", "小6"],
            },
            "junior_boys": {
                "level": "中学校",
                "gender": "男性",
                "grades": ["中1", "中2", "中3"],
            },
            "junior_girls": {
                "level": "中学校",
                "gender": "女性",
                "grades": ["中1", "中2", "中3"],
            },
        }

        if series_type not in series_def:
            raise ValueError(f"不明な survey_series: {series_type}")

        defn = series_def[series_type]
        level = defn["level"]
        gender = defn["gender"]
        grades = defn["grades"]

        # フィルタ列の存在確認
        for col in ["school_level", "gender_norm", "grade_2024"]:
            if col not in df.columns:
                raise RuntimeError(f"必要な列 '{col}' が見つかりません。Excelや前処理の仕様をご確認ください。")

        mask_level_gender = (df["school_level"] == level) & (df["gender_norm"] == gender)
        counts: List[int] = []
        for g in grades:
            counts.append(int(((df["grade_2024"] == g) & mask_level_gender).sum()))
        return counts

    except Exception as e:
        raise RuntimeError(f"シリーズ集計の取得に失敗しました: {e}")


# クリーム色（変更時）と薄い水色（未変更時）の塗りつぶし
CREAM_FILL = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
BLUE_FILL = PatternFill(fill_type="solid", start_color="DDEBF7", end_color="DDEBF7")


def write_with_cream(ws, addr: str, new_value: Any):
    """セルの値を設定し、
    - 値が変わった場合: クリーム色（FFF2CC）
    - 値が変わらない場合: 薄い水色（DDEBF7）
    の背景を付与する。
    """
    try:
        old_value = ws[addr].value
    except Exception:
        old_value = None
    ws[addr] = new_value
    try:
        if old_value != new_value:
            ws[addr].fill = CREAM_FILL
        else:
            ws[addr].fill = BLUE_FILL
    except Exception:
        # スタイル設定に失敗しても処理を継続
        pass


def fill_from_yaml(config_path: Union[str, Path]) -> Path:
    """
    YAML 設定を読み込み、指定のシート/セルへ値を書き込みます。

    想定する YAML 例:

    template: report_template.xlsx
    output: report_result.xlsx
    survey_excel: survey.xlsx  # オプション：アンケートExcelファイルのパス
    writes:
      - sheet: p1
        cell: AC14
        value: xxx
      - sheet: p2
        row: 5
        column: B   # または 2 （数値でも可）
        value: Hello
      - sheet: p1
        cell: B10
        survey_data: total_responses      # 総回答者数
      - sheet: p1
        cell: B11
        survey_data: invalid_responses    # 無効回答者数
      - sheet: p1
        cell: B12
        survey_data: effective_responses  # 有効回答者数

      # 新機能: 連続セルに学年別の人数を書き込み（デフォルトは下方向）
      - sheet: p1
        cell: D20      # このセルに小1男子、その下に小2男子…小6男子まで
        survey_series: elementary_boys   # elementary_boys | elementary_girls | junior_boys | junior_girls
        # direction: down  # 省略可（right も指定可能）

      # 新機能: 地域別（小1〜中3）を横方向に書き込み（デフォルトは右方向）
      - sheet: p1
        cell: E25
        survey_series: region_grades
        region: 東京23区  # 許可: 東京23区 / 三多摩島しょ / 埼玉県 / 神奈川県 / 千葉県 / その他
        # direction: right  # 省略可（未指定なら右）
        # または、region_grades:東京23区 のように survey_series に地域名を含めても可

      # 追加機能: 指定設問の選択肢の回答数（学年または地域ごと）を縦方向に書き込み
      - sheet: Q1
        cell: T10
        survey_series: responses
        question: 1          # main.py の設問一覧順（1開始）
        choices:
          - 知っている       # 複数指定可／1件のみでも可
          - 知らない
        class: grade         # grade | region
        # 備考: choices を複数指定すると、各選択肢の列を右方向に展開し、各列は縦方向（down）に書き込みます

    :param config_path: YAML ファイルへのパス
    :return: 出力されたファイルの Path
    """
    if yaml is None:
        raise RuntimeError("PyYAML が必要です。`pip install pyyaml` を実行してください。")

    cpath = Path(config_path)
    if not cpath.exists():
        raise FileNotFoundError(f"設定ファイルが見つかりません: {cpath}")

    with cpath.open("r", encoding="utf-8") as f:
        data = yaml.safe_load(f) or {}

    if not isinstance(data, dict):
        raise ValueError("YAML のルートはマッピングである必要があります（dict）。")

    # 必須キー
    template = data.get("template")
    output = data.get("output")
    writes = data.get("writes")
    survey_excel = data.get("survey_excel", "survey.xlsx")  # デフォルトは survey.xlsx

    if not template:
        raise ValueError("YAML に 'template' が必要です。")
    if not output:
        raise ValueError("YAML に 'output' が必要です。")
    if not isinstance(writes, list) or not writes:
        raise ValueError("YAML の 'writes' は1件以上のリストである必要があります。")

    # パスは設定ファイルの場所からの相対も許可
    base = cpath.parent
    tpath = (base / template) if not Path(str(template)).is_absolute() else Path(str(template))
    out = (base / output) if not Path(str(output)).is_absolute() else Path(str(output))
    survey_path = (base / survey_excel) if not Path(str(survey_excel)).is_absolute() else Path(str(survey_excel))

    if not tpath.exists():
        raise FileNotFoundError(f"テンプレートが見つかりません: {tpath}")

    # パフォーマンス最適化: データ処理を最初に1回だけ実行
    processed_data = None
    needs_survey_data = any(
        w.get("survey_data") is not None or w.get("survey_series") is not None or w.get("survey_pick_count") is not None
        for w in writes
    )
    if needs_survey_data and ReportDataPreparator is not None:
        try:
            config = ReportConfig()
            preparator = ReportDataPreparator(config)
            processed_data = preparator.prepare_data(survey_path)
        except Exception as e:
            print(f"警告: survey データの事前読み込みに失敗しました: {e}")
            processed_data = None

    wb = load_workbook(tpath)

    for i, w in enumerate(writes, start=1):
        if not isinstance(w, dict):
            raise ValueError(f"writes[{i}] はマッピングである必要があります。")
        sheet = w.get("sheet")
        if not sheet:
            raise ValueError(f"writes[{i}] に 'sheet' がありません。")
        if sheet not in wb.sheetnames:
            raise KeyError(f"シートが見つかりません: '{sheet}'（writes[{i}]）")

        # アドレスの決定: cell 優先、なければ row+column
        addr = w.get("cell")
        if not addr:
            row = w.get("row")
            col = w.get("column")
            if row is None or col is None:
                raise ValueError(f"writes[{i}] には 'cell' か 'row'+'column' のいずれかが必要です。")
            # column は文字列（例: 'AC'）または数値（例: 29）を許可
            if isinstance(col, int):
                col_letter = get_column_letter(col)
            else:
                col_letter = str(col).strip()
                if not col_letter:
                    raise ValueError(f"writes[{i}] の column が不正です。")
            addr = f"{col_letter}{int(row)}"

        # シリーズ書き込みか単一値かを判定
        series_type = w.get("survey_series")
        value = w.get("value")
        survey_data_type = w.get("survey_data")

        # survey_pick_count の処理
        pick_count = w.get("survey_pick_count")
        
        # 相互排他（survey_pick_count を追加）
        specified = [x is not None for x in (value, survey_data_type, series_type, pick_count)]
        if sum(specified) != 1:
            raise ValueError(f"writes[{i}] では 'value' / 'survey_data' / 'survey_series' / 'survey_pick_count' のいずれか1つだけを指定してください。")

        ws = wb[sheet]

        if series_type is not None:
            # region_grades の地域名を補完（survey_series が文字列 'region_grades' かつ region キーが与えられている場合）
            region_param = w.get("region")
            if isinstance(series_type, str) and series_type.strip() == "region_grades":
                if region_param:
                    series_type = f"region_grades:{str(region_param).strip()}"
                else:
                    raise ValueError(f"writes[{i}] の region_grades には 'region' キーで地域名を指定してください（例: region: 東京23区）。")

            # responses / ratios の補完（question / choices / class を付与）
            # - choices（リスト）を推奨。choice（単一）は後方互換としてサポート（非推奨）。
            is_responses = isinstance(series_type, str) and series_type.strip() == "responses"
            is_ratios = isinstance(series_type, str) and series_type.strip() == "ratios"
            choices_list = None
            if is_responses or is_ratios:
                q = w.get("question")
                cls = w.get("class")
                # 新仕様: 複数選択肢
                if "choices" in w and w.get("choices") is not None:
                    cl = w.get("choices")
                    if not isinstance(cl, list) or not cl:
                        raise ValueError(f"writes[{i}] の responses: 'choices' は1件以上のリストで指定してください。")
                    choices_list = [str(x).strip() for x in cl]
                else:
                    # 従来の 'choice' を後方互換でサポート（非推奨）
                    ch = w.get("choice")
                    if ch is not None:
                        choices_list = [str(ch).strip()]
                    else:
                        raise ValueError(f"writes[{i}] の {'responses' if is_responses else 'ratios'} には 'choices'（リスト）を指定してください。旧 'choice' は廃止予定です。")
                if q is None or cls is None:
                    raise ValueError(f"writes[{i}] の {'responses' if is_responses else 'ratios'} には question / class を指定してください。")
                try:
                    q_int = int(q)
                except Exception:
                    raise ValueError(f"writes[{i}] の question は整数で指定してください（指定値: {q}）。")
                # series_type は choices ごとに後で組み立てる

            # 連続セルにシリーズを書き込み
            # 方向のデフォルト: 通常は down、region_grades は right
            default_direction = "right" if isinstance(series_type, str) and isinstance(series_type, str) and series_type.startswith("region_grades") else "down"
            direction = (w.get("direction") or default_direction).lower()
            if direction not in ("down", "right"):
                raise ValueError(f"writes[{i}] の direction は 'down' または 'right' で指定してください。")

            # アドレス分解
            col_letters = ''.join([c for c in addr if c.isalpha()])
            row_digits = ''.join([c for c in addr if c.isdigit()])
            if not col_letters or not row_digits:
                raise ValueError(f"writes[{i}] の cell アドレスが不正です: {addr}")
            base_col = col_letters
            base_row = int(row_digits)

            # 列文字<->番号の相互変換
            def col_to_num(s: str) -> int:
                num = 0
                for ch in s:
                    num = num * 26 + (ord(ch.upper()) - ord('A') + 1)
                return num
            def num_to_col(n: int) -> str:
                res = ""
                while n > 0:
                    n, rem = divmod(n - 1, 26)
                    res = chr(ord('A') + rem) + res
                return res
            base_col_num = col_to_num(base_col)

            if is_ratios and choices_list is not None:
                # ratios 機能: 複数 choices の割合を横並びで出力
                cls = str(w.get("class")).strip()
                choices_str = ",".join(choices_list)
                series_str = f"ratios:q={q_int};choices={choices_str};class={cls}"
                try:
                    # choice_mappingを取得（YAMLから）
                    yaml_choice_mapping = w.get("choice_mapping", {})
                    yaml_select_count = w.get("select_count")
                    values = get_survey_data_series(series_str, processed_data, survey_path, yaml_choice_mapping, choices_list, yaml_select_count)
                except Exception as e:
                    raise RuntimeError(f"writes[{i}] の ratios 取得に失敗: {e}")
                
                # direction に応じて配置
                if direction == "right":
                    # 横方向に配置（選択肢ごとに列をずらす）
                    for idx, v in enumerate(values):
                        col_letter = num_to_col(base_col_num + idx)
                        target = f"{col_letter}{base_row}"
                        write_with_cream(ws, target, v)
                else:
                    # 縦方向に配置
                    for idx, v in enumerate(values):
                        target = f"{base_col}{base_row + idx}"
                        write_with_cream(ws, target, v)
            elif is_responses and choices_list is not None:
                # 新仕様: 複数 choices に対応
                # - 各 choice のシリーズは「縦方向（down）」に書き込み
                # - choice ごとに開始列を1つずつ右にずらす
                cls = str(w.get("class")).strip()
                for j, ch_text in enumerate(choices_list):
                    series_str = f"responses:q={q_int};choice={ch_text};class={cls}"
                    try:
                        # choice_mappingを取得（YAMLから）
                        yaml_choice_mapping = w.get("choice_mapping", {})
                        yaml_select_count = w.get("select_count")
                        values = get_survey_data_series(series_str, processed_data, survey_path, yaml_choice_mapping, choices_list, yaml_select_count)
                    except Exception as e:
                        raise RuntimeError(f"writes[{i}] の responses 取得に失敗（choice='{ch_text}'）: {e}")
                    # 常に縦方向に配置（仕様: 複数 choices は列方向に展開）
                    col_letter = num_to_col(base_col_num + j)
                    for idx, v in enumerate(values):
                        target = f"{col_letter}{base_row + idx}"
                        write_with_cream(ws, target, v)
            else:
                # 従来通りの単一シリーズ書き込み
                try:
                    # choice_mappingを取得（YAMLから）
                    yaml_choice_mapping = w.get("choice_mapping", {})
                    yaml_select_count = w.get("select_count")
                    values = get_survey_data_series(series_type, processed_data, survey_path, yaml_choice_mapping, choices_list, yaml_select_count)
                except Exception as e:
                    raise RuntimeError(f"writes[{i}] の survey_series '{series_type}' の取得に失敗: {e}")

                if direction == "down":
                    for idx, v in enumerate(values):
                        target = f"{base_col}{base_row + idx}"
                        write_with_cream(ws, target, v)
                else:  # right
                    for idx, v in enumerate(values):
                        col_letter = num_to_col(base_col_num + idx)
                        target = f"{col_letter}{base_row}"
                        write_with_cream(ws, target, v)
        elif pick_count is not None:
            # survey_pick_count の処理
            q = w.get("question")
            if q is None:
                raise ValueError(f"writes[{i}] の survey_pick_count には question を指定してください。")
            try:
                q_int = int(q)
                count_int = int(pick_count)
            except Exception:
                raise ValueError(f"writes[{i}] の question と survey_pick_count は整数で指定してください（question: {q}, survey_pick_count: {pick_count}）。")
            
            pick_count_str = f"pick_count:q={q_int};count={count_int}"
            try:
                final_value = get_survey_data_value(pick_count_str, processed_data, survey_path)
            except Exception as e:
                raise RuntimeError(f"writes[{i}] の survey_pick_count の取得に失敗: {e}")
            write_with_cream(ws, addr, final_value)
        else:
            # 単一セル書き込み
            if value is not None:
                final_value = value
            else:
                try:
                    # survey_data: multiple の場合は追加パラメータを渡す
                    if survey_data_type == "multiple":
                        q = w.get("question")
                        choices_list = w.get("choices")
                        yaml_choice_mapping = w.get("choice_mapping")
                        cls = w.get("class")
                        
                        if q is None or not choices_list:
                            raise ValueError(f"writes[{i}] の survey_data: multiple には question と choices が必要です。")
                        
                        try:
                            q_int = int(q)
                        except Exception:
                            raise ValueError(f"writes[{i}] の question は整数で指定してください（指定値: {q}）。")
                        
                        final_value = get_survey_data_value(survey_data_type, processed_data, survey_path, 
                                                          question=q_int, choices=choices_list, 
                                                          choice_mapping=yaml_choice_mapping, class_type=cls)
                        
                        # class が grade/region の場合は配列が返されるので、連続セルに書き込み
                        if cls and cls.lower() in ("grade", "region") and isinstance(final_value, list):
                            # アドレス分解
                            col_letters = ''.join([c for c in addr if c.isalpha()])
                            row_digits = ''.join([c for c in addr if c.isdigit()])
                            if not col_letters or not row_digits:
                                raise ValueError(f"writes[{i}] の cell アドレスが不正です: {addr}")
                            base_col = col_letters
                            base_row = int(row_digits)
                            
                            # 方向のデフォルト: down
                            direction = (w.get("direction") or "down").lower()
                            if direction not in ("down", "right"):
                                raise ValueError(f"writes[{i}] の direction は 'down' または 'right' で指定してください。")
                            
                            # 列文字<->番号の相互変換
                            def col_to_num(s: str) -> int:
                                num = 0
                                for ch in s:
                                    num = num * 26 + (ord(ch.upper()) - ord('A') + 1)
                                return num
                            def num_to_col(n: int) -> str:
                                res = ""
                                while n > 0:
                                    n, rem = divmod(n - 1, 26)
                                    res = chr(ord('A') + rem) + res
                                return res
                            base_col_num = col_to_num(base_col)
                            
                            # 配列の各要素を書き込み
                            if direction == "down":
                                for idx, v in enumerate(final_value):
                                    target = f"{base_col}{base_row + idx}"
                                    write_with_cream(ws, target, v)
                            else:  # right
                                for idx, v in enumerate(final_value):
                                    col_letter = num_to_col(base_col_num + idx)
                                    target = f"{col_letter}{base_row}"
                                    write_with_cream(ws, target, v)
                            continue  # 単一セル書き込みはスキップ
                        
                    else:
                        final_value = get_survey_data_value(survey_data_type, processed_data, survey_path)
                except Exception as e:
                    raise RuntimeError(f"writes[{i}] の survey_data '{survey_data_type}' の取得に失敗: {e}")
            
            # 単一セル書き込み（配列書き込みでcontinueされた場合はここは実行されない）
            write_with_cream(ws, addr, final_value)

    # 出力ディレクトリが存在しない場合に備える
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def main():
    """
    YAML 設定で一括書き込みを実行します（デフォルト）。

    使い方:
      python fill_template_excel.py <config.yaml>

    例:
      python fill_template_excel.py apps/report_sgk/sample_fill.yaml

    備考:
      以前の「引数なしで p1!AC14 に 'xxx' を書き込む」動作は削除されました。
    """
    if len(sys.argv) < 2:
        print("エラー: YAML 設定ファイルへのパスを指定してください。\n"
              "使い方: python fill_template_excel.py <config.yaml>\n"
              "例:     python fill_template_excel.py apps/report_sgk/sample_fill.yaml")
        sys.exit(2)

    cfg = Path(sys.argv[1])
    out_path = fill_from_yaml(cfg)
    print(f"YAML 設定に従い書き出し完了: {out_path}")


if __name__ == "__main__":
    main()
